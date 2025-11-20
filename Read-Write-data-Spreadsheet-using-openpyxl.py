import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "EmployeeData"

sheet['A1'] = "Name"
sheet['B1'] = "Department"
sheet['C1'] = "Salary"

sheet.append(["sasi", "HR", 30000])
sheet.append(["ravi", "IT", 40000])
sheet.append(["anil", "Finance", 35000])

wb.save("employee_data.xlsx")

wb2 = openpyxl.load_workbook("employee_data.xlsx")
sheet2 = wb2.active

for row in sheet2.iter_rows(values_only=True):
    print(row)

